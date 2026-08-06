from __future__ import annotations

import re
from collections import Counter
from io import BytesIO
from typing import Any
from urllib.parse import quote_plus, urlencode

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app import models
from app.database import get_db

router = APIRouter(prefix="/component-catalog", tags=["Catálogo de componentes"])

TYPE_MAP = {
    "qf": "interruptor", "qs": "seccionador", "gv": "guardamotor",
    "km": "contactor", "ka": "relé", "fr": "relé térmico", "fu": "fusible",
    "vfd": "variador", "uf": "variador", "atv": "variador",
    "plc": "PLC", "di": "módulo de entradas", "do": "módulo de salidas",
    "ai": "módulo analógico", "ao": "módulo analógico", "m": "motor",
    "b": "sensor", "x": "bornera", "xt": "bornera", "h": "piloto",
    "s": "pulsador", "t": "transformador",
}


MODEL_TYPE_RULES = (
    (re.compile(r"^3RV", re.IGNORECASE), "guardamotor"),
    (re.compile(r"^3RT", re.IGNORECASE), "contactor"),
    (re.compile(r"^(ATV|ALTIVAR)", re.IGNORECASE), "variador"),
    (re.compile(r"^(FC-|FC\s)?\d{2,4}", re.IGNORECASE), "variador"),
    (re.compile(r"^6ES7", re.IGNORECASE), "módulo PLC"),
)

MODEL_MANUFACTURER_RULES = (
    (re.compile(r"^(3RV|3RT|6ES7|6SL)", re.IGNORECASE), "Siemens"),
    (re.compile(r"^(ATV|LC1D|GV2|TM3|BMX)", re.IGNORECASE), "Schneider Electric"),
    (re.compile(r"^(FC-|VLT)", re.IGNORECASE), "Danfoss"),
    (re.compile(r"^(ACS|AF|MS116)", re.IGNORECASE), "ABB"),
)



NON_COMPONENT_EXACT = {
    "PE", "N", "L", "L1", "L2", "L3", "L+", "L-", "+24V", "24V", "0V",
    "M", "SH", "SHIELD", "SCHIRM",
    "BN", "BK", "BU", "BL", "GY", "GN", "YE", "GNYE", "GN/YE", "RD", "WH", "OG", "VT",
    "ST", "PIEZA", "RES",
}
NON_COMPONENT_PATTERNS = (
    re.compile(r"^(?:L[123]|PE|N|0V|24V|\+24V|L\+|L-)$", re.IGNORECASE),
    re.compile(r"^(?:BN|BK|BU|BL|GY|GN|YE|GNYE|GN/YE|RD|WH|OG|VT)$", re.IGNORECASE),
)

def is_nonphysical_reference(reference: str | None) -> bool:
    """True para potenciales, colores y etiquetas que no representan un componente físico."""
    raw = (reference or "").strip().upper()
    compact = re.sub(r"\s+", "", raw)
    if compact in NON_COMPONENT_EXACT:
        return True
    return any(pattern.fullmatch(compact) for pattern in NON_COMPONENT_PATTERNS)



PLC_CHANNEL_RE = re.compile(
    r"^(?:[IQM](?:[WDB])?\d+(?:\.\d+)?|[AQ](?:[WDB])?\d+(?:\.\d+)?)$",
    re.IGNORECASE,
)

def is_plc_channel(value: str | None) -> bool:
    return bool(PLC_CHANNEL_RE.fullmatch((value or "").strip()))

def _term_location_payload(term: models.PageSearchTerm, page: models.DocumentPage, document: models.Document) -> dict[str, Any]:
    return {
        "document_id": document.id,
        "document_title": document.title,
        "page_id": page.id,
        "page_number": page.page_number,
        "x": term.x,
        "y": term.y,
        "width": term.width,
        "height": term.height,
        "display_text": term.display_text or term.term or "",
        "row_text": term.row_text or "",
    }


def _reference_location_payload(ref: models.ComponentReference, page: models.DocumentPage, document: models.Document) -> dict[str, Any]:
    return {
        "reference_id": ref.id,
        "document_id": document.id,
        "document_title": document.title,
        "page_id": page.id,
        "page_number": page.page_number,
        "x": ref.x,
        "y": ref.y,
        "width": ref.width,
        "height": ref.height,
        "display_text": ref.reference or "",
        "row_text": ref.row_text or ref.description or "",
        "model": ref.model or "",
        "source_kind": ref.source_kind or "",
        "page_type": page.page_type or "",
    }


def _plc_channel_occurrences(db: Session, item: dict[str, Any], query_value: str) -> list[dict[str, Any]]:
    """Todas las apariciones exactas del canal dentro del documento del módulo."""
    wanted = normalize_term(query_value)
    document_id = item.get("document_id")
    if not wanted or not document_id:
        return []
    rows = (
        db.query(models.PageSearchTerm, models.DocumentPage, models.Document)
        .join(models.DocumentPage, models.PageSearchTerm.document_page_id == models.DocumentPage.id)
        .join(models.Document, models.DocumentPage.document_id == models.Document.id)
        .filter(models.Document.id == document_id)
        .all()
    )
    found: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for term, page, document in rows:
        candidates = (getattr(term, "term", None), getattr(term, "display_text", None))
        if not any(normalize_term(candidate) == wanted for candidate in candidates):
            continue
        key = (page.id, term.x, term.y, term.width, term.height)
        if key in seen:
            continue
        seen.add(key)
        found.append(_term_location_payload(term, page, document))
    return sorted(found, key=lambda row: (row["page_number"], row.get("y") or 0, row.get("x") or 0))


def _module_occurrence_score(row: dict[str, Any], reference: str, model: str) -> tuple[int, int, int, int]:
    text = f"{row.get('row_text') or ''} {row.get('display_text') or ''}".upper()
    page_type = (row.get("page_type") or "").lower()
    source = (row.get("source_kind") or "").lower()
    score = 0
    if model and normalize_term(model) in normalize_term(text):
        score += 8
    if reference and normalize_term(reference) in normalize_term(text):
        score += 5
    if source in {"schematic", "plan", "catalog+plan"}:
        score += 3
    if page_type in {"schematic", "plan", "electrical", "diagram"}:
        score += 3
    if row.get("x") is not None and row.get("y") is not None:
        score += 2
    # Las tablas de asignación de direcciones son útiles para el canal, pero no
    # deben ser la ubicación física principal del módulo.
    if any(token in text for token in ("DIRECCIÓN", "DIRECTION", "ADRESSE", "DIGITALES-AUSGANGSMODUL", "DIGITALES-EINGANGSMODUL")):
        score -= 5
    return (score, 1 if model and normalize_term(model) in normalize_term(text) else 0, 1 if row.get("x") is not None else 0, -(row.get("page_number") or 0))


def _infer_module_instance_reference(db: Session, channel_occurrences: list[dict[str, Any]], channel: str) -> str:
    """Intenta obtener la designación física del módulo (p. ej. KE1.3) desde la página del canal."""
    channel_norm = normalize_term(channel)
    for occurrence in channel_occurrences[:5]:
        page_id = occurrence.get("page_id")
        if not page_id:
            continue
        page = db.query(models.DocumentPage).filter(models.DocumentPage.id == page_id).first()
        if not page or not page.text_content:
            continue
        text = page.text_content
        compact = normalize_term(text)
        pos = compact.find(channel_norm)
        # Si la normalización impide localizar el índice exacto, usamos toda la página.
        window = text if pos < 0 else text[max(0, pos - 1200): pos + 1200]
        tags = re.findall(r"[-=]([A-Z]{1,5}\d+(?:\.\d+)*)", window.upper())
        # Priorizamos designaciones típicas de módulos/racks y descartamos el propio canal.
        ranked = []
        for tag in tags:
            if normalize_term(tag) == channel_norm:
                continue
            score = 0
            if re.match(r"^(?:KE|ET|PLC|A|E|K)\d", tag):
                score += 5
            if "." in tag:
                score += 2
            if tag.startswith(("Q", "I")):
                score -= 3
            ranked.append((score, tag))
        if ranked:
            ranked.sort(reverse=True)
            return ranked[0][1]
    return ""


def _plc_module_occurrences(
    db: Session,
    item: dict[str, Any],
    module_instance_reference: str = "",
) -> list[dict[str, Any]]:
    """Ubicaciones físicas plausibles del módulo, sin búsquedas masivas por textos genéricos."""
    document_id = item.get("document_id")
    reference = (item.get("reference") or item.get("parent_reference") or "").strip()
    model = (item.get("model") or "").strip()
    if not document_id:
        return []

    rows = (
        db.query(models.ComponentReference, models.DocumentPage, models.Document)
        .join(models.DocumentPage, models.ComponentReference.document_page_id == models.DocumentPage.id)
        .join(models.Document, models.DocumentPage.document_id == models.Document.id)
        .filter(models.Document.id == document_id)
        .all()
    )

    candidates: list[dict[str, Any]] = []
    seen_pages: set[int] = set()
    ref_norm = normalize_term(reference)
    model_norm = normalize_term(model)
    instance_norm = normalize_term(module_instance_reference)

    for ref, page, document in rows:
        page_text = page.text_content or ""
        row_text = f"{ref.row_text or ''} {ref.description or ''} {ref.reference or ''} {ref.model or ''}"
        page_norm = normalize_term(page_text)
        row_norm = normalize_term(row_text)
        ref_model = _best_model_from_page(ref.reference or "", page_text, ref.model)
        same_model = bool(model_norm and normalize_term(ref_model) == model_norm)
        exact_model_in_row = bool(model_norm and model_norm in row_norm)
        exact_model_on_page = bool(model_norm and model_norm in page_norm)
        same_reference = bool(ref_norm and normalize_term(ref.reference) == ref_norm)
        instance_on_page = bool(instance_norm and instance_norm in page_norm)
        instance_in_row = bool(instance_norm and instance_norm in row_norm)

        # Con una designación física inferida, exigimos que la página contenga esa
        # designación y el modelo exacto. Así evitamos cientos/miles de páginas del
        # mismo modelo instaladas en otros racks o sectores.
        if instance_norm:
            if not ((instance_on_page or instance_in_row) and (exact_model_in_row or exact_model_on_page or same_model)):
                continue
        else:
            # Sin designación física, solo aceptamos coincidencias fuertes: modelo
            # exacto junto al registro, o referencia exacta acompañada por el modelo.
            if not (exact_model_in_row or (same_reference and exact_model_on_page)):
                continue

        if page.id in seen_pages:
            continue
        seen_pages.add(page.id)
        payload = _reference_location_payload(ref, page, document)
        payload["model"] = ref_model or model
        payload["module_instance_reference"] = module_instance_reference
        candidates.append(payload)

    candidates.sort(key=lambda row: _module_occurrence_score(row, module_instance_reference or reference, model), reverse=True)
    # Una ficha física debe mostrar pocas ubicaciones candidatas y verificables.
    return candidates[:20]

def _attach_plc_channel_target(db: Session, item: dict[str, Any], query_value: str) -> bool:
    """Vincula canal y módulo sin confundir sus ubicaciones.

    La tarjeta representa el canal consultado, pero conserva dos conjuntos de
    ubicaciones: las apariciones exactas del canal y las del hardware físico.
    """
    channel_occurrences = _plc_channel_occurrences(db, item, query_value)
    if not channel_occurrences:
        return False
    module_instance_reference = _infer_module_instance_reference(db, channel_occurrences, query_value)
    module_occurrences = _plc_module_occurrences(db, item, module_instance_reference)

    parent = item.get("reference") or "módulo"
    channel = query_value.strip().upper()
    channel_primary = channel_occurrences[0]
    module_primary = module_occurrences[0] if module_occurrences else None

    item["parent_reference"] = parent
    item["channel_reference"] = channel
    item["display_reference"] = channel
    item["component_type"] = "canal PLC"
    item["match_rank"] = 0
    item["match_reason"] = f"Canal exacto del módulo {parent}"
    item["channel_occurrences"] = channel_occurrences
    item["module_occurrences"] = module_occurrences
    item["module_instance_reference"] = module_instance_reference
    item["channel_page_count"] = len({row["page_number"] for row in channel_occurrences})
    item["module_page_count"] = len({row["page_number"] for row in module_occurrences})

    # La ubicación principal de la ficha es la del módulo físico si se pudo
    # determinar; la del canal queda separada y se abre con su propio botón.
    primary = module_primary or channel_primary
    for field in ("document_id", "document_title", "page_id", "page_number", "x", "y", "width", "height"):
        item[field] = primary.get(field)
    item["channel_document_id"] = channel_primary.get("document_id")
    item["channel_page_id"] = channel_primary.get("page_id")
    item["channel_page_number"] = channel_primary.get("page_number")
    item["channel_x"] = channel_primary.get("x")
    item["channel_y"] = channel_primary.get("y")
    item["channel_width"] = channel_primary.get("width")
    item["channel_height"] = channel_primary.get("height")
    item["description"] = (
        f"Canal {channel} perteneciente al módulo {parent}. "
        + (item.get("description") or "")
    ).strip()
    return True

MANUFACTURER_DOMAINS = {
    "siemens": "siemens.com",
    "schneider": "se.com",
    "schneider electric": "se.com",
    "danfoss": "danfoss.com",
    "abb": "abb.com",
    "sew": "sew-eurodrive.com",
    "sew-eurodrive": "sew-eurodrive.com",
    "rockwell": "rockwellautomation.com",
    "allen-bradley": "rockwellautomation.com",
    "omron": "omron.com",
    "sick": "sick.com",
    "ifm": "ifm.com",
    "festo": "festo.com",
    "wago": "wago.com",
    "phoenix": "phoenixcontact.com",
    "weidmuller": "weidmueller.com",
    "weidmüller": "weidmueller.com",
    "eaton": "eaton.com",
}


def infer_manufacturer(model: str | None, manufacturer: str | None) -> str:
    current = (manufacturer or "").strip()
    if current:
        return current
    clean_model = (model or "").strip()
    for pattern, name in MODEL_MANUFACTURER_RULES:
        if pattern.search(clean_model):
            return name
    return ""


def _model_type(model: str | None) -> str:
    clean_model = (model or "").strip()
    for pattern, component_type in MODEL_TYPE_RULES:
        if pattern.search(clean_model):
            return component_type
    return ""


def official_component_links(manufacturer: str | None, model: str | None) -> dict[str, str]:
    """Genera búsquedas acotadas al sitio oficial sin inventar una URL de producto."""
    maker = (manufacturer or "").strip()
    part = (model or "").strip()
    if not part:
        return {"product_url": "", "manual_url": ""}
    domain = MANUFACTURER_DOMAINS.get(maker.lower())
    if not domain:
        return {"product_url": "", "manual_url": ""}
    product_query = quote_plus(f"site:{domain} {part}")
    manual_query = quote_plus(f"site:{domain} {part} manual OR datasheet PDF")
    return {
        "product_url": f"https://www.google.com/search?q={product_query}",
        "manual_url": f"https://www.google.com/search?q={manual_query}",
    }


def infer_type(reference: str, detected: str | None, component_type: str | None, model: str | None = None, description: str | None = None) -> str:
    model_type = _model_type(model)
    if model_type:
        return model_type
    context = " ".join(filter(None, (detected, component_type, description))).strip().lower()
    for keyword, label in (
        ("guardamotor", "guardamotor"),
        ("protector de motor", "guardamotor"),
        ("contactor", "contactor"),
        ("relé térmico", "relé térmico"),
        ("rele termico", "relé térmico"),
        ("variador", "variador"),
        ("motor", "motor"),
        ("sensor", "sensor"),
        ("fin de carrera", "contacto o fin de carrera"),
    ):
        if keyword in context:
            return label
    for value in (detected, component_type):
        if value and value.strip() and value.strip().lower() not in {"referencia técnica", "referencia fc"}:
            return value.strip().lower()
    ref = (reference or "").strip().lower()
    prefix = "".join(ch for ch in ref if ch.isalpha())
    if prefix.startswith("fc"):
        return "referencia FC"
    for key in sorted(TYPE_MAP, key=len, reverse=True):
        if prefix.startswith(key):
            return TYPE_MAP[key]
    return "otro"


def normalize_term(value: str | None) -> str:
    """Normaliza referencias para que fc011, -FC011 y =DV2-FC011 coincidan."""
    text = (value or "").upper().strip()
    text = re.sub(r"\s+", "", text)
    # Conserva letras y números; elimina separadores de plano.
    return re.sub(r"[^A-Z0-9]", "", text)


RELIABLE_MODEL_RE = re.compile(
    r"\b(?:3RV\d+[A-Z0-9-]*|3RT\d+[A-Z0-9-]*|6ES7[A-Z0-9-]+|6SL[A-Z0-9-]+|"
    r"ATV[A-Z0-9-]+|LC1D[A-Z0-9-]+|GV2[A-Z0-9-]+|VLT[A-Z0-9-]+|ACS[A-Z0-9-]+|"
    r"MS116[A-Z0-9-]*|AF\d+[A-Z0-9-]*)\b",
    re.IGNORECASE,
)


def _is_reliable_model(value: str | None) -> bool:
    return bool(RELIABLE_MODEL_RE.fullmatch((value or "").strip()))


def _best_model_from_page(reference: str, page_text: str | None, current_model: str | None) -> str:
    """Recupera el modelo real desde el texto cercano del plano.

    Evita aceptar otra referencia (por ejemplo FC061) como modelo de FC011.
    """
    current = (current_model or "").strip()
    if _is_reliable_model(current):
        return current
    text = page_text or ""
    if not text:
        return ""
    candidates = list(RELIABLE_MODEL_RE.finditer(text))
    if not candidates:
        return ""
    ref = (reference or "").strip()
    ref_positions = [m.start() for m in re.finditer(re.escape(ref), text, re.IGNORECASE)] if ref else []
    if not ref_positions:
        # Si la página tiene un solo modelo técnico inequívoco, se puede usar.
        unique = []
        seen = set()
        for match in candidates:
            value = match.group(0).upper()
            if value not in seen:
                seen.add(value)
                unique.append(match.group(0))
        return unique[0] if len(unique) == 1 else ""
    best = min(
        candidates,
        key=lambda match: min(abs(match.start() - pos) for pos in ref_positions),
    )
    distance = min(abs(best.start() - pos) for pos in ref_positions)
    # Ventana suficientemente amplia para rótulos verticales, sin mezclar toda la página.
    return best.group(0) if distance <= 500 else ""


def _description_score(value: str | None) -> int:
    text = (value or "").strip()
    if not text:
        return 0
    score = min(len(text), 240)
    for keyword in ("TRANSPORT", "MOTOR", "BOMBA", "VENTIL", "IZQUIER", "DERECH", "LUBRIC"):
        if keyword in text.upper():
            score += 80
    return score


def _base_query(
    db: Session,
    organization_id: int | None,
    plant_id: int | None,
    sector_id: int | None,
):
    query = (
        db.query(
            models.ComponentReference,
            models.DocumentPage,
            models.Document,
            models.Sector,
            models.Plant,
            models.Organization,
        )
        .join(models.DocumentPage, models.ComponentReference.document_page_id == models.DocumentPage.id)
        .join(models.Document, models.DocumentPage.document_id == models.Document.id)
        .join(models.Sector, models.Document.sector_id == models.Sector.id)
        .join(models.Plant, models.Sector.plant_id == models.Plant.id)
        .join(models.Organization, models.Plant.organization_id == models.Organization.id)
    )
    if organization_id is not None:
        query = query.filter(models.Organization.id == organization_id)
    if plant_id is not None:
        query = query.filter(models.Plant.id == plant_id)
    if sector_id is not None:
        query = query.filter(models.Sector.id == sector_id)
    return query


def _row_to_item(row: tuple[Any, ...], search_normalized: str = "") -> dict[str, Any]:
    ref, page, doc, sector, plant, org = row
    reference = ref.reference or ""
    row_text = ref.row_text or ""
    model = _best_model_from_page(reference, page.text_content, ref.model)
    normalized_reference = normalize_term(reference)
    normalized_model = normalize_term(model)

    match_rank = 99
    match_reason = ""
    if search_normalized:
        if normalized_reference == search_normalized:
            match_rank, match_reason = 0, "Coincidencia exacta en referencia"
        elif normalized_model == search_normalized:
            match_rank, match_reason = 1, "Coincidencia exacta en modelo"
        elif search_normalized and search_normalized in normalized_reference:
            match_rank, match_reason = 2, "Coincidencia parcial en referencia"
        elif search_normalized and search_normalized in normalized_model:
            match_rank, match_reason = 3, "Coincidencia parcial en modelo"
        elif search_normalized and search_normalized in normalize_term(row_text):
            match_rank, match_reason = 4, "Mencionado en la descripción"

    manufacturer = infer_manufacturer(model, getattr(ref, "manufacturer", None))
    component_type = infer_type(reference, ref.detected_type, ref.component_type, model, ref.description or row_text)
    links = official_component_links(manufacturer, model)

    return {
        "id": ref.id,
        "reference": reference,
        "display_reference": reference,
        "parent_reference": "",
        "channel_reference": "",
        "component_type": component_type,
        "model": model,
        "manufacturer": manufacturer,
        "product_url": links["product_url"],
        "manual_url": links["manual_url"],
        "source_kind": getattr(ref, "source_kind", None) or "",
        "catalog_confidence": getattr(ref, "catalog_confidence", 0) or 0,
        "description": ref.description or row_text,
        "document_id": doc.id,
        "document_title": doc.title,
        "page_number": page.page_number,
        "page_id": page.id,
        "x": ref.x, "y": ref.y, "width": ref.width, "height": ref.height,
        "organization_id": org.id, "organization_name": org.name,
        "plant_id": plant.id, "plant_name": plant.name,
        "sector_id": sector.id, "sector_name": sector.name,
        "match_rank": match_rank,
        "match_reason": match_reason,
    }


PHYSICAL_LIBRARY_TYPES = {
    "guardamotor", "contactor", "interruptor", "seccionador", "relé", "relé térmico",
    "fusible", "variador", "PLC", "módulo PLC", "módulo de entradas",
    "módulo de salidas", "módulo analógico", "motor", "sensor", "válvula",
    "fuente", "transformador", "bornera", "pulsador", "piloto",
}

CHANNEL_OR_SUBELEMENT_RE = re.compile(
    r"^(?:AI|AO|DI|DO|DQ|IQ|Q|I|IW|QW|IB|QB|ID|QD)\d+(?:[_.]\d+)?$",
    re.IGNORECASE,
)

def is_library_equipment(item: dict[str, Any], include_incomplete: bool = False) -> bool:
    """Deja en Biblioteca solo equipos físicos; señales y canales quedan en Buscar."""
    reference = (item.get("reference") or "").strip()
    if is_nonphysical_reference(reference) or CHANNEL_OR_SUBELEMENT_RE.fullmatch(reference):
        return False
    model = (item.get("model") or "").strip()
    component_type = (item.get("component_type") or "").strip()
    reliable = _is_reliable_model(model)
    physical_type = component_type in PHYSICAL_LIBRARY_TYPES
    if include_incomplete:
        return reliable or physical_type
    return reliable and physical_type


def _filtered_items(
    db: Session,
    organization_id: int | None,
    plant_id: int | None,
    sector_id: int | None,
    component_type: str | None,
    q: str | None,
    hard_limit: int,
) -> list[dict[str, Any]]:
    query = _base_query(db, organization_id, plant_id, sector_id)
    search_normalized = normalize_term(q)

    # Reduce candidatos en SQL, pero la prioridad final se calcula normalizada en Python.
    if q and q.strip():
        raw = q.strip()
        term = f"%{raw}%"
        normalized_term = f"%{search_normalized}%"
        query = query.filter(
            models.ComponentReference.reference.ilike(term)
            | models.ComponentReference.reference.ilike(normalized_term)
            | models.ComponentReference.model.ilike(term)
            | models.ComponentReference.model.ilike(normalized_term)
            | models.ComponentReference.row_text.ilike(term)
            | models.ComponentReference.normalized_reference.ilike(normalized_term)
        )

    rows = query.order_by(
        models.ComponentReference.reference.asc(),
        models.DocumentPage.page_number.asc(),
    ).limit(hard_limit).all()

    items = [_row_to_item(row, search_normalized) for row in rows]
    # La biblioteca contiene solo componentes físicos. Potenciales y colores siguen
    # disponibles en Buscar, pero no generan fichas ni filas de Excel.
    items = [item for item in items if not is_nonphysical_reference(item.get("reference"))]
    if component_type:
        wanted = component_type.strip().lower()
        items = [item for item in items if item["component_type"].lower() == wanted]

    if search_normalized:
        # Si existen coincidencias en referencia/modelo, se eliminan menciones secundarias.
        best_rank = min((item["match_rank"] for item in items), default=99)
        if best_rank <= 3:
            items = [item for item in items if item["match_rank"] <= 3]
        items.sort(key=lambda item: (
            item["match_rank"],
            normalize_term(item["reference"]),
            item["page_number"],
        ))
    return items


@router.get("")
def list_components(
    organization_id: int | None = Query(default=None),
    plant_id: int | None = Query(default=None),
    sector_id: int | None = Query(default=None),
    component_type: str | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=2000),
    include_incomplete: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    raw_items = _filtered_items(
        db, organization_id, plant_id, sector_id, component_type, q, hard_limit=10000
    )
    raw_items = [item for item in raw_items if is_library_equipment(item, include_incomplete)]

    # Consolida apariciones repetidas: una ficha por referencia real, documento y sector.
    # Las menciones en listas, contactos auxiliares y referencias cruzadas quedan agrupadas.
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    occurrences: dict[tuple[Any, ...], list[dict[str, Any]]] = {}

    def quality(item: dict[str, Any]) -> tuple[int, int, int, int, int, int]:
        source = (item.get("source_kind") or "").lower()
        model = item.get("model") or ""
        return (
            1 if _is_reliable_model(model) else 0,
            1 if source in {"schematic", "plan", "catalog+plan"} else 0,
            1 if item.get("manufacturer") else 0,
            int(item.get("catalog_confidence") or 0),
            _description_score(item.get("description")),
            1 if item.get("x") is not None and item.get("y") is not None else 0,
        )

    for item in raw_items:
        key = (
            normalize_term(item.get("reference")),
            item.get("document_id"),
            item.get("sector_id"),
        )
        occurrences.setdefault(key, []).append(item)
        current = grouped.get(key)
        if current is None or quality(item) > quality(current):
            grouped[key] = dict(item)

    consolidated = []
    for key, item in grouped.items():
        rows = occurrences[key]
        pages = sorted({int(row["page_number"]) for row in rows if row.get("page_number") is not None})
        item["occurrence_count"] = len(rows)
        item["page_count"] = len(pages)
        item["occurrence_pages"] = pages[:50]
        # Completa tipo y fabricante después de recuperar el mejor modelo de plano.
        item["manufacturer"] = infer_manufacturer(item.get("model"), item.get("manufacturer"))
        item["component_type"] = infer_type(
            item.get("reference") or "", item.get("component_type"), item.get("component_type"),
            item.get("model"), item.get("description"),
        )
        item.update(official_component_links(item.get("manufacturer"), item.get("model")))
        consolidated.append(item)

    # Para direcciones PLC (Q5050.0, I123.4, QW5260, etc.) no se muestra
    # la etiqueta genérica del módulo (M0). Se conserva la ficha del hardware,
    # pero la tarjeta y la navegación apuntan al canal exacto indexado.
    if q and is_plc_channel(q):
        exact_channel_items = []
        for item in consolidated:
            candidate = dict(item)
            if _attach_plc_channel_target(db, candidate, q):
                exact_channel_items.append(candidate)
        if exact_channel_items:
            consolidated = exact_channel_items

    consolidated.sort(key=lambda item: (
        item.get("match_rank", 99),
        normalize_term(item.get("display_reference") or item.get("reference")),
        item.get("sector_name") or "",
        item.get("document_title") or "",
    ))
    consolidated = consolidated[:limit]
    counts = Counter(item["component_type"] for item in consolidated)
    return {"items": consolidated, "counts": dict(sorted(counts.items())), "total": len(consolidated)}


@router.get("/export.xlsx")
def export_components_excel(
    organization_id: int | None = Query(default=None),
    plant_id: int | None = Query(default=None),
    sector_id: int | None = Query(default=None),
    component_type: str | None = Query(default=None),
    q: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """Exporta una hoja técnica consolidada y otra con todas las apariciones."""
    items = _filtered_items(
        db, organization_id, plant_id, sector_id, component_type, q, hard_limit=50000
    )

    def quality(item: dict[str, Any]) -> tuple[int, int, int, int, int]:
        model = (item.get("model") or "").upper()
        reliable_model = bool(re.match(r"^(3RV|3RT|6ES7|6SL|ATV|LC1D|GV2|FC-|VLT|ACS|AF|MS116)", model))
        source = (item.get("source_kind") or "").lower()
        component_type_value = (item.get("component_type") or "").lower()
        return (
            1 if reliable_model else 0,
            1 if source == "catalog+plan" else 0,
            1 if item.get("manufacturer") else 0,
            int(item.get("catalog_confidence") or 0),
            1 if component_type_value not in {"otro", "referencia técnica", "referencia fc"} else 0,
        )

    # Una ficha por referencia/modelo/sector. Si hay varias apariciones, conserva la más rica.
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for item in items:
        reliable_model = normalize_term(item.get("model"))
        key = (
            normalize_term(item.get("reference")),
            reliable_model if reliable_model else "SINMODELO",
            item.get("sector_id"),
        )
        previous = grouped.get(key)
        if previous is None or quality(item) > quality(previous):
            grouped[key] = item
    unique = sorted(grouped.values(), key=lambda x: (
        normalize_term(x.get("reference")), x.get("sector_name") or "", x.get("page_number") or 0
    ))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Componentes únicos"
    headers = [
        "Referencia", "Tipo", "Fabricante", "Modelo completo", "Función / descripción",
        "Empresa", "Planta", "Sector", "Documento", "Página principal",
        "Confianza", "Origen", "Página oficial", "Manual / ficha técnica",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item in unique:
        row = [
            item["reference"], item["component_type"], item["manufacturer"],
            item["model"], item["description"], item["organization_name"],
            item["plant_name"], item["sector_name"], item["document_title"],
            item["page_number"], item["catalog_confidence"], item["source_kind"],
            "Abrir producto oficial" if item.get("product_url") else "",
            "Buscar manual oficial" if item.get("manual_url") else "",
        ]
        sheet.append(row)
        row_index = sheet.max_row
        if item.get("product_url"):
            sheet.cell(row_index, 13).hyperlink = item["product_url"]
            sheet.cell(row_index, 13).style = "Hyperlink"
        if item.get("manual_url"):
            sheet.cell(row_index, 14).hyperlink = item["manual_url"]
            sheet.cell(row_index, 14).style = "Hyperlink"

    occurrences = workbook.create_sheet("Apariciones")
    occurrence_headers = [
        "Referencia", "Tipo", "Fabricante", "Modelo", "Empresa", "Planta", "Sector",
        "Documento", "Página", "Origen", "Descripción",
    ]
    occurrences.append(occurrence_headers)
    for cell in occurrences[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    seen_occurrences: set[tuple[Any, ...]] = set()
    for item in items:
        key = (item["id"], item["document_id"], item["page_number"])
        if key in seen_occurrences:
            continue
        seen_occurrences.add(key)
        occurrences.append([
            item["reference"], item["component_type"], item["manufacturer"], item["model"],
            item["organization_name"], item["plant_name"], item["sector_name"],
            item["document_title"], item["page_number"], item["source_kind"], item["description"],
        ])

    for current in (sheet, occurrences):
        current.freeze_panes = "A2"
        current.auto_filter.ref = current.dimensions
        for row in current.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [18, 24, 22, 30, 55, 22, 22, 25, 35, 14, 12, 18, 24, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    occurrence_widths = [18, 24, 22, 30, 22, 22, 25, 35, 10, 18, 60]
    for index, width in enumerate(occurrence_widths, start=1):
        occurrences.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="diagramiq-componentes.xlsx"'},
    )

